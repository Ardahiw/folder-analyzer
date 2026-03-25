import os
import csv
import shutil
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from makeModelClassificationDataOps import DataFolderAnalyzer


class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None

        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event=None):
        if self.tipwindow or not self.text:
            return

        x = self.widget.winfo_rootx() + self.widget.winfo_width() + 10
        y = self.widget.winfo_rooty() + 8

        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")

        label = tk.Label(
            tw,
            text=self.text,
            justify="left",
            background="#111827",
            foreground="#f8fafc",
            relief="solid",
            borderwidth=1,
            font=("Segoe UI", 9),
            padx=8,
            pady=4
        )
        label.pack()

    def hide_tooltip(self, event=None):
        if self.tipwindow:
            self.tipwindow.destroy()
            self.tipwindow = None


class FolderAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Klasör Analiz Aracı")
        self.root.geometry("1250x780")
        self.root.minsize(1050, 700)

        self.selected_folder = ""
        self.last_result = None
        self.selected_folders = set()
        self.valid_folders = set()

        self.sidebar_expanded = True
        self.sidebar_expanded_width = 240
        self.sidebar_collapsed_width = 78

        self.current_active = None

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.colors = {
            "bg_sidebar": "#111827",
            "bg_main": "#0f172a",
            "surface": "#1e293b",
            "surface_2": "#243041",
            "surface_3": "#2b3648",
            "border": "#334155",
            "text": "#f8fafc",
            "text_muted": "#94a3b8",
            "primary": "#1d4ed8",
            "primary_hover": "#2563eb",
            "accent": "#f59e0b",
            "accent_hover": "#fbbf24",
            "danger": "#b91c1c",
            "danger_hover": "#dc2626",
            "success": "#166534",
            "warning": "#92400e",
            "info": "#0f3d73",
            "purple": "#4c1d95",
        }

        self.is_busy = False
        self.search_var = tk.StringVar()
        self.low_threshold_var = tk.StringVar(value="5")
        self.current_filtered_data = []

        self.sort_column = None
        self.sort_reverse = False
        self._loading_text_job = None
        self._loading_frame_job = None
        self._loading_spinner_index = 0
        self._loading_message_base = "Analiz hazırlanıyor"
        self._sidebar_animating = False
        self.preview_thumbnails = []
        self.preview_image_names = []

        self.build_ui()

    def build_ui(self):
        self.root.configure(fg_color=self.colors["bg_main"])

        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=0, minsize=self.sidebar_expanded_width)
        self.root.grid_columnconfigure(1, weight=1)

        # ================= SIDEBAR =================
        self.sidebar_frame = ctk.CTkFrame(
            self.root,
            width=self.sidebar_expanded_width,
            corner_radius=0,
            fg_color=self.colors["bg_sidebar"]
        )
        self.sidebar_frame.grid(row=0, column=0, sticky="nsw")
        self.sidebar_frame.grid_columnconfigure(0, weight=1)
        self.sidebar_frame.grid_columnconfigure(1, weight=0)
        self.sidebar_frame.grid_rowconfigure(15, weight=1)

        # ================= MAIN =================
        self.main_frame = ctk.CTkFrame(
            self.root,
            corner_radius=18,
            fg_color=self.colors["bg_main"]
        )
        self.main_frame.grid(row=0, column=1, padx=(0, 18), pady=18, sticky="nsew")
        self.main_frame.grid_rowconfigure(0, weight=0)
        self.main_frame.grid_rowconfigure(1, weight=0)
        self.main_frame.grid_rowconfigure(2, weight=0)
        self.main_frame.grid_rowconfigure(4, weight=0)
        self.main_frame.grid_rowconfigure(5, weight=3)
        self.main_frame.grid_rowconfigure(6, weight=0)
        self.main_frame.grid_columnconfigure(0, weight=1)

        # ================= SIDEBAR HEADER =================
        self.logo_label = ctk.CTkLabel(
            self.sidebar_frame,
            text="🧠 Analyzer",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=24, weight="bold")
        )
        self.logo_label.grid(row=0, column=0, padx=24, pady=(24, 8), sticky="w")

        self.toggle_button = ctk.CTkButton(
            self.sidebar_frame,
            text="☰",
            width=40,
            height=40,
            command=self.toggle_sidebar,
            corner_radius=10,
            fg_color=self.colors["surface_2"],
            hover_color=self.colors["surface_3"],
            text_color=self.colors["text"],
            border_width=1,
            border_color=self.colors["border"],
            cursor="hand2"
        )
        self.toggle_button.grid(row=0, column=1, padx=14, pady=(24, 8), sticky="e")

        self.sidebar_subtitle = ctk.CTkLabel(
            self.sidebar_frame,
            text="Klasör analiz ve yönetim paneli",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=12),
            justify="left"
        )
        self.sidebar_subtitle.grid(row=1, column=0, padx=24, pady=(0, 20), sticky="w")

        self.nav_title = ctk.CTkLabel(
            self.sidebar_frame,
            text="MENÜ",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.nav_title.grid(row=2, column=0, padx=24, pady=(10, 8), sticky="w")

        button_style = {
            "height": 42,
            "corner_radius": 12,
            "anchor": "w",
            "font": ctk.CTkFont(size=13, weight="bold"),
            "fg_color": self.colors["surface_2"],
            "hover_color": self.colors["surface_3"],
            "text_color": self.colors["text"],
            "border_width": 1,
            "border_color": self.colors["border"],
            "cursor": "hand2"
        }

        self.sidebar_select_btn = ctk.CTkButton(
            self.sidebar_frame,
            text="📂 Klasör Seç",
            command=self.select_folder,
            **button_style
        )
        self.sidebar_select_btn.grid(row=3, column=0, columnspan=2, padx=14, pady=6, sticky="ew")

        self.sidebar_analyze_btn = ctk.CTkButton(
            self.sidebar_frame,
            text="🔍 Analiz Et",
            command=self.analyze_selected_folder,
            **button_style
        )
        self.sidebar_analyze_btn.grid(row=4, column=0, columnspan=2, padx=14, pady=6, sticky="ew")

        self.sidebar_open_btn = ctk.CTkButton(
            self.sidebar_frame,
            text="📁 Seçiliyi Aç",
            command=self.open_selected_folder,
            **button_style
        )
        self.sidebar_open_btn.grid(row=5, column=0, columnspan=2, padx=14, pady=6, sticky="ew")

        self.sidebar_move_btn = ctk.CTkButton(
            self.sidebar_frame,
            text="🚚 Klasörü Taşı",
            command=self.move_selected_folder,
            **button_style
        )
        self.sidebar_move_btn.grid(row=6, column=0, columnspan=2, padx=14, pady=6, sticky="ew")

        self.sidebar_health_btn = ctk.CTkButton(
            self.sidebar_frame,
            text="🩺 Sağlık Kontrolü",
            command=self.run_health_check,
            **button_style
        )
        self.sidebar_health_btn.grid(row=7, column=0, columnspan=2, padx=14, pady=6, sticky="ew")

        self.sidebar_select_low_btn = ctk.CTkButton(
            self.sidebar_frame,
            text="⚡ Az Verili Seç",
            command=self.select_low_folders,
            **button_style
        )
        self.sidebar_select_low_btn.grid(row=8, column=0, columnspan=2, padx=14, pady=6, sticky="ew")

        self.sidebar_bulk_move_btn = ctk.CTkButton(
            self.sidebar_frame,
            text="🚀 Seçilenleri Taşı",
            command=self.move_selected_folders_bulk,
            **button_style
        )
        self.sidebar_bulk_move_btn.grid(row=9, column=0, columnspan=2, padx=14, pady=6, sticky="ew")

        self.export_title = ctk.CTkLabel(
            self.sidebar_frame,
            text="DIŞA AKTAR",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.export_title.grid(row=10, column=0, padx=24, pady=(18, 8), sticky="w")

        self.sidebar_csv_btn = ctk.CTkButton(
            self.sidebar_frame,
            text="🧾 CSV Kaydet",
            command=self.export_to_csv,
            **button_style
        )
        self.sidebar_csv_btn.grid(row=11, column=0, columnspan=2, padx=14, pady=6, sticky="ew")

        self.sidebar_excel_btn = ctk.CTkButton(
            self.sidebar_frame,
            text="📊 Excel Kaydet",
            command=self.export_to_excel,
            **button_style
        )
        self.sidebar_excel_btn.grid(row=12, column=0, columnspan=2, padx=14, pady=6, sticky="ew")

        self.actions_title = ctk.CTkLabel(
            self.sidebar_frame,
            text="İŞLEMLER",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.actions_title.grid(row=13, column=0, padx=24, pady=(18, 8), sticky="w")

        self.sidebar_clear_btn = ctk.CTkButton(
            self.sidebar_frame,
            text="🧹 Tablo Temizle",
            command=self.clear_list,
            height=42,
            anchor="w",
            corner_radius=12,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#3b1f24",
            hover_color="#4a2329",
            text_color=self.colors["text"],
            border_width=1,
            border_color="#6b2c35",
            cursor="hand2"
        )
        self.sidebar_clear_btn.grid(row=14, column=0, columnspan=2, padx=14, pady=6, sticky="ew")

        self.sidebar_info = ctk.CTkLabel(
            self.sidebar_frame,
            text="Beklemede\nBir klasör seçip analize başlayabilirsin.",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=12),
            justify="left"
        )
        self.sidebar_info.grid(row=15, column=0, padx=24, pady=(20, 24), sticky="sw")

        # ================= HEADER =================
        self.header_frame = ctk.CTkFrame(
            self.main_frame,
            corner_radius=16,
            fg_color=self.colors["surface"]
        )
        self.header_frame.grid(row=0, column=0, padx=16, pady=(16, 10), sticky="ew")
        self.header_frame.grid_columnconfigure(0, weight=1)
        self.header_frame.grid_columnconfigure(1, weight=0)

        self.title_label = ctk.CTkLabel(
            self.header_frame,
            text="🗂️ Klasör Analiz Aracı",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=28, weight="bold")
        )
        self.title_label.grid(row=0, column=0, padx=18, pady=(14, 4), sticky="w")

        self.desc_label = ctk.CTkLabel(
            self.header_frame,
            text="Veri klasörlerini analiz et, kategorilere ayır, filtrele ve dışa aktar.",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=13)
        )
        self.desc_label.grid(row=1, column=0, padx=18, pady=(0, 14), sticky="w")

        self.status_badge = ctk.CTkLabel(
            self.header_frame,
            text="Beklemede",
            width=120,
            corner_radius=12,
            fg_color="#1e293b",
            text_color="#e2e8f0",
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.status_badge.grid(row=0, column=1, padx=(8, 18), pady=(14, 6), sticky="e")

        self.refresh_button = ctk.CTkButton(
            self.header_frame,
            text="↻",
            width=36,
            height=36,
            command=self.refresh_analysis,
            corner_radius=10,
            fg_color=self.colors["surface_2"],
            hover_color=self.colors["surface_3"],
            text_color=self.colors["text"],
            border_width=1,
            border_color=self.colors["border"],
            font=ctk.CTkFont(size=16, weight="bold"),
            cursor="hand2"
        )
        self.refresh_button.grid(row=1, column=1, padx=(8, 18), pady=(0, 14), sticky="e")

        self.progress = ctk.CTkProgressBar(
            self.header_frame,
            height=8,
            corner_radius=999,
            progress_color=self.colors["accent"],
            fg_color=self.colors["surface_2"]
        )
        self.progress.grid(row=2, column=0, columnspan=2, padx=18, pady=(0, 14), sticky="ew")
        self.progress.set(0)
        self.progress.grid_remove()

        # ================= PATH =================
        self.path_frame = ctk.CTkFrame(
            self.main_frame,
            corner_radius=16,
            fg_color=self.colors["surface"]
        )
        self.path_frame.grid(row=1, column=0, padx=16, pady=(0, 10), sticky="ew")
        self.path_frame.grid_columnconfigure(0, weight=1)

        self.path_title = ctk.CTkLabel(
            self.path_frame,
            text="Seçilen Klasör",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=13, weight="bold")
        )
        self.path_title.grid(row=0, column=0, padx=16, pady=(10, 2), sticky="w")

        self.folder_info_var = tk.StringVar(value="Henüz klasör seçilmedi.")
        self.folder_info_label = ctk.CTkLabel(
            self.path_frame,
            textvariable=self.folder_info_var,
            text_color=self.colors["text_muted"],
            anchor="w",
            justify="left",
            wraplength=1100,
            font=ctk.CTkFont(size=12)
        )
        self.folder_info_label.grid(row=1, column=0, padx=16, pady=(0, 8), sticky="ew")

        # ===== ANALİZ MODU =====
        self.analysis_mode_label = ctk.CTkLabel(
            self.path_frame,
            text="Analiz Modu",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=13, weight="bold")
        )
        self.analysis_mode_label.grid(row=2, column=0, padx=16, pady=(4, 2), sticky="w")

        self.analysis_mode_row = ctk.CTkFrame(
            self.path_frame,
            fg_color="transparent"
        )
        self.analysis_mode_row.grid(row=3, column=0, padx=16, pady=(0, 10), sticky="w")

        self.analysis_mode = ctk.CTkSegmentedButton(
            self.analysis_mode_row,
            values=["Standart", "Recursive"],
            fg_color=self.colors["surface_2"],
            selected_color=self.colors["accent"],
            selected_hover_color=self.colors["accent_hover"],
            unselected_color=self.colors["surface_2"],
            unselected_hover_color=self.colors["surface_3"],
            text_color="#111827",
            font=ctk.CTkFont(size=12, weight="bold"),
            width=220
        )
        self.analysis_mode.pack(side="left")
        self.analysis_mode.set("Recursive")
        self.analysis_mode.configure(command=self.on_analysis_mode_change)

        self.analysis_mode_info = ctk.CTkLabel(
            self.analysis_mode_row,
            text="Seçilen klasörün alt klasörleri dahil tüm içeriğini gösterir.",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=11),
            justify="left"
        )
        self.analysis_mode_info.pack(side="left", padx=(12, 0))

        # ================= STATS =================
        self.stats_frame = ctk.CTkFrame(
            self.main_frame,
            corner_radius=16,
            fg_color=self.colors["bg_main"]
        )
        self.stats_frame.grid(row=2, column=0, padx=16, pady=(0, 10), sticky="ew")

        for i in range(4):
            self.stats_frame.grid_columnconfigure(i, weight=1)

        self.total_folders_card = ctk.CTkFrame(
            self.stats_frame,
            corner_radius=14,
            fg_color="#1f2937",
            border_width=1,
            border_color="#334155"
        )
        self.total_folders_card.grid(row=0, column=0, padx=8, pady=10, sticky="ew")

        self.total_images_card = ctk.CTkFrame(
            self.stats_frame,
            corner_radius=14,
            fg_color="#1f2937",
            border_width=1,
            border_color="#334155"
        )
        self.total_images_card.grid(row=0, column=1, padx=8, pady=10, sticky="ew")

        self.low_count_card = ctk.CTkFrame(
            self.stats_frame,
            corner_radius=14,
            fg_color="#1f2937",
            border_width=1,
            border_color="#4b5563"
        )
        self.low_count_card.grid(row=0, column=2, padx=8, pady=10, sticky="ew")

        self.high_count_card = ctk.CTkFrame(
            self.stats_frame,
            corner_radius=14,
            fg_color="#1f2937",
            border_width=1,
            border_color="#4b5563"
        )
        self.high_count_card.grid(row=0, column=3, padx=8, pady=10, sticky="ew")

        self.bind_card_hover(self.total_folders_card, "#1f2937", "#273449")
        self.bind_card_hover(self.total_images_card, "#1f2937", "#2d304a")
        self.bind_card_hover(self.low_count_card, "#1f2937", "#3a2a2a")
        self.bind_card_hover(self.high_count_card, "#1f2937", "#23352d")

        self.total_folders_title = ctk.CTkLabel(
            self.total_folders_card,
            text="📁 Toplam Klasör",
            text_color="#93c5fd",
            font=ctk.CTkFont(size=13, weight="bold")
        )
        self.total_folders_title.pack(anchor="w", padx=14, pady=(10, 2))

        self.total_folders_value = ctk.CTkLabel(
            self.total_folders_card,
            text="0",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=22, weight="bold")
        )
        self.total_folders_value.pack(anchor="w", padx=14, pady=(0, 10))

        self.total_images_title = ctk.CTkLabel(
            self.total_images_card,
            text="🖼️ Toplam Görüntü",
            text_color="#c4b5fd",
            font=ctk.CTkFont(size=13, weight="bold")
        )
        self.total_images_title.pack(anchor="w", padx=14, pady=(10, 2))

        self.total_images_value = ctk.CTkLabel(
            self.total_images_card,
            text="0",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=22, weight="bold")
        )
        self.total_images_value.pack(anchor="w", padx=14, pady=(0, 10))

        self.low_count_title = ctk.CTkLabel(
            self.low_count_card,
            text="🔴 0-5 Kategorisi",
            text_color="#fca5a5",
            font=ctk.CTkFont(size=13, weight="bold")
        )
        self.low_count_title.pack(anchor="w", padx=14, pady=(10, 2))

        self.low_count_value = ctk.CTkLabel(
            self.low_count_card,
            text="0",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=22, weight="bold")
        )
        self.low_count_value.pack(anchor="w", padx=14, pady=(0, 10))

        self.high_count_title = ctk.CTkLabel(
            self.high_count_card,
            text="🟢 50+ Kategorisi",
            text_color="#86efac",
            font=ctk.CTkFont(size=13, weight="bold")
        )
        self.high_count_title.pack(anchor="w", padx=14, pady=(10, 2))

        self.high_count_value = ctk.CTkLabel(
            self.high_count_card,
            text="0",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=22, weight="bold")
        )
        self.high_count_value.pack(anchor="w", padx=14, pady=(0, 10))

        # ================= FILTER =================
        self.filter_frame = ctk.CTkFrame(
            self.main_frame,
            corner_radius=16,
            fg_color=self.colors["surface"]
        )
        self.filter_frame.grid(row=4, column=0, padx=16, pady=(0, 10), sticky="ew")
        self.filter_frame.grid_columnconfigure(4, weight=1)

        self.filter_label = ctk.CTkLabel(
            self.filter_frame,
            text="Kategori Filtresi",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=14, weight="bold")
        )
        self.filter_label.grid(row=0, column=0, padx=(16, 10), pady=12, sticky="w")

        self.filter_segmented = ctk.CTkSegmentedButton(
            self.filter_frame,
            values=["Tümü", "0-5", "6-25", "26-50", "50+"],
            command=self.on_filter_change,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color=self.colors["surface_2"],
            selected_color=self.colors["accent"],
            selected_hover_color=self.colors["accent_hover"],
            unselected_color=self.colors["surface_2"],
            unselected_hover_color=self.colors["surface_3"],
            text_color="#111827",
            text_color_disabled=self.colors["text_muted"]
        )
        self.filter_segmented.grid(row=0, column=1, padx=10, pady=12, sticky="w")
        self.filter_segmented.set("Tümü")

        self.threshold_label = ctk.CTkLabel(
            self.filter_frame,
            text="Threshold",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=12, weight="bold")
        )
        self.threshold_label.grid(row=0, column=2, padx=(12, 6), pady=12, sticky="w")

        self.threshold_entry = ctk.CTkEntry(
            self.filter_frame,
            textvariable=self.low_threshold_var,
            width=60,
            height=34,
            corner_radius=10,
            fg_color=self.colors["surface_2"],
            border_color=self.colors["border"],
            text_color=self.colors["text"],
            justify="center"
        )
        self.threshold_entry.grid(row=0, column=3, padx=(0, 8), pady=12, sticky="w")

        self.search_container = ctk.CTkFrame(
            self.filter_frame,
            fg_color="transparent"
        )
        self.search_container.grid(row=0, column=4, padx=(10, 16), pady=12, sticky="e")

        self.search_label = ctk.CTkLabel(
            self.search_container,
            text="🔎 Arama",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=13, weight="bold")
        )
        self.search_label.pack(side="left", padx=(0, 6))

        self.search_entry = ctk.CTkEntry(
            self.search_container,
            textvariable=self.search_var,
            placeholder_text="Klasör ara...",
            width=220,
            height=36,
            corner_radius=10,
            fg_color=self.colors["surface_2"],
            border_color=self.colors["border"],
            text_color=self.colors["text"]
        )
        self.search_entry.pack(side="left")
        self.search_entry.bind("<KeyRelease>", self.on_search_change)
        self.search_entry.bind(
            "<FocusIn>",
            lambda e: self.search_entry.configure(border_color=self.colors["accent"])
        )
        self.search_entry.bind(
            "<FocusOut>",
            lambda e: self.search_entry.configure(border_color=self.colors["border"])
        )

        self.search_clear_btn = ctk.CTkButton(
            self.search_container,
            text="✕",
            width=30,
            height=30,
            corner_radius=8,
            fg_color=self.colors["surface_2"],
            hover_color="#7f1d1d",
            text_color=self.colors["text_muted"],
            border_width=1,
            border_color=self.colors["border"],
            command=self.clear_search,
            cursor="hand2"
        )
        self.search_clear_btn.pack(side="left", padx=(6, 0))

        # ================= TABLE CARD =================
        self.table_card = ctk.CTkFrame(
            self.main_frame,
            corner_radius=16,
            fg_color=self.colors["surface"]
        )
        self.table_card.grid(row=5, column=0, padx=16, pady=(0, 8), sticky="nsew")
        self.table_card.grid_rowconfigure(1, weight=1)
        self.table_card.grid_columnconfigure(0, weight=3)
        self.table_card.grid_columnconfigure(1, weight=2)

        self.summary_var = tk.StringVar(value="Durum: Beklemede")
        self.summary_label = ctk.CTkLabel(
            self.table_card,
            textvariable=self.summary_var,
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=13, weight="bold")
        )
        self.summary_label.grid(row=0, column=0, padx=16, pady=(14, 8), sticky="w")

        self.table_frame = tk.Frame(
            self.table_card,
            bg="#1e293b",
            bd=0,
            highlightthickness=0
        )
        self.table_frame.grid(row=1, column=0, padx=(16, 8), pady=(0, 12), sticky="nsew")
        self.table_frame.grid_rowconfigure(0, weight=1)
        self.table_frame.grid_columnconfigure(0, weight=1)

        style = ttk.Style()
        style.theme_use("clam")

        style.configure(
            "Treeview",
            background="#1e293b",
            foreground="white",
            fieldbackground="#1e293b",
            rowheight=36,
            font=("Segoe UI", 10),
            borderwidth=0
        )

        style.configure(
            "Treeview.Heading",
            background="#243041",
            foreground="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat"
        )

        style.map(
            "Treeview",
            background=[("selected", "#2563eb")],
            foreground=[("selected", "white")]
        )

        style.map(
            "Treeview.Heading",
            background=[("active", "#2b3648"), ("pressed", "#243041")],
            foreground=[("active", "white"), ("pressed", "white")]
        )

        self.tree = ttk.Treeview(
            self.table_frame,
            columns=("folder", "count", "category"),
            show="headings"
        )

        self.tree.heading("folder", text="Klasör", command=lambda: self.sort_treeview("folder"))
        self.tree.heading("count", text="Görüntü Sayısı", command=lambda: self.sort_treeview("count"))
        self.tree.heading("category", text="Kategori", command=lambda: self.sort_treeview("category"))

        self.tree.column("folder", width=700, anchor="w")
        self.tree.column("count", width=170, anchor="center")
        self.tree.column("category", width=170, anchor="center")

        self.tree.grid(row=0, column=0, sticky="nsew")

        self.scrollbar = ttk.Scrollbar(
            self.table_frame,
            orient="vertical",
            command=self.tree.yview
        )
        self.scrollbar.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=self.scrollbar.set)

        self.tree.bind("<Double-1>", self.open_selected_folder)
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.tree.bind("<Motion>", self.update_tree_cursor)
        self.tree.bind("<Leave>", lambda e: self.tree.configure(cursor=""))

        self.preview_frame = ctk.CTkFrame(
            self.table_card,
            corner_radius=14,
            fg_color=self.colors["surface_2"],
            border_width=1,
            border_color=self.colors["border"]
        )
        self.preview_frame.grid(row=1, column=1, padx=(8, 16), pady=(0, 12), sticky="nsew")
        self.preview_frame.grid_columnconfigure(0, weight=1)
        self.preview_frame.grid_rowconfigure(2, weight=1)

        self.preview_title = ctk.CTkLabel(
            self.preview_frame,
            text="🖼️ Klasör Önizleme",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=18, weight="bold")
        )
        self.preview_title.grid(row=0, column=0, padx=14, pady=(14, 6), sticky="w")

        self.preview_info = ctk.CTkLabel(
            self.preview_frame,
            text="Tablodan bir klasör seç.\nBurada örnek görseller ve hızlı özet görünecek.",
            text_color=self.colors["text_muted"],
            justify="left",
            anchor="w",
            font=ctk.CTkFont(size=12)
        )
        self.preview_info.grid(row=1, column=0, padx=14, pady=(0, 10), sticky="ew")

        self.preview_images_frame = ctk.CTkFrame(
            self.preview_frame,
            fg_color="transparent"
        )
        self.preview_images_frame.grid(row=2, column=0, padx=10, pady=(0, 10), sticky="nsew")

        for i in range(3):
            self.preview_images_frame.grid_columnconfigure(i, weight=1)

        self.preview_hint = ctk.CTkLabel(
            self.preview_images_frame,
            text="Önizleme için klasör seçilmedi",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=13, weight="bold")
        )
        self.preview_hint.grid(row=0, column=0, columnspan=3, padx=8, pady=16)

        self.preview_footer = ctk.CTkLabel(
            self.preview_frame,
            text="İlk 6 görsel gösterilir",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=11)
        )
        self.preview_footer.grid(row=3, column=0, padx=14, pady=(0, 12), sticky="w")

        self.empty_placeholder = ctk.CTkLabel(
            self.table_card,
            text="Henüz analiz sonucu yok.\nBir klasör seçip analiz başlatabilirsin.",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=15, weight="bold"),
            justify="center"
        )
        self.empty_placeholder.grid(row=1, column=0, sticky="")

        self.loading_overlay = ctk.CTkFrame(
            self.table_card,
            corner_radius=16,
            fg_color="#0b1120"
        )
        self.loading_overlay.grid(row=0, column=0, rowspan=2, columnspan=2, padx=10, pady=10, sticky="nsew")
        self.loading_overlay.grid_columnconfigure(0, weight=1)
        self.loading_overlay.grid_rowconfigure((0, 1, 2, 3), weight=1)

        self.loading_spinner_label = ctk.CTkLabel(
            self.loading_overlay,
            text="◐",
            text_color=self.colors["accent"],
            font=ctk.CTkFont(size=40, weight="bold")
        )
        self.loading_spinner_label.grid(row=0, column=0, pady=(40, 6))

        self.loading_title_label = ctk.CTkLabel(
            self.loading_overlay,
            text="Analiz sürüyor",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=22, weight="bold")
        )
        self.loading_title_label.grid(row=1, column=0, pady=(0, 4))

        self.loading_message_var = tk.StringVar(value="Klasörler hazırlanıyor...")
        self.loading_message_label = ctk.CTkLabel(
            self.loading_overlay,
            textvariable=self.loading_message_var,
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=13),
            justify="center"
        )
        self.loading_message_label.grid(row=2, column=0, padx=24, pady=(0, 4))

        self.loading_detail_var = tk.StringVar(value="0 / 0 klasör işlendi")
        self.loading_detail_label = ctk.CTkLabel(
            self.loading_overlay,
            textvariable=self.loading_detail_var,
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=12)
        )
        self.loading_detail_label.grid(row=3, column=0, pady=(0, 28))
        self.loading_overlay.grid_remove()

        # ================= TOOLTIPS =================
        self.sidebar_select_tooltip = ToolTip(self.sidebar_select_btn, "Klasör Seç")
        self.sidebar_analyze_tooltip = ToolTip(self.sidebar_analyze_btn, "Analiz Et")
        self.sidebar_open_tooltip = ToolTip(self.sidebar_open_btn, "Seçiliyi Aç")
        self.sidebar_move_tooltip = ToolTip(self.sidebar_move_btn, "Klasörü Taşı")
        self.sidebar_health_tooltip = ToolTip(self.sidebar_health_btn, "Dataset sağlık kontrolü yap")
        self.sidebar_select_low_tooltip = ToolTip(self.sidebar_select_low_btn, "Seçilen threshold değerine uygun olan klasörleri seç")
        self.sidebar_bulk_move_tooltip = ToolTip(self.sidebar_bulk_move_btn, "Seçilen az verili klasörleri toplu taşı")
        self.sidebar_csv_tooltip = ToolTip(self.sidebar_csv_btn, "Analizi CSV olarak kaydet (.csv)")
        self.sidebar_excel_tooltip = ToolTip(self.sidebar_excel_btn, "Analizi Excel Kaydet (.xlsx)")
        self.sidebar_clear_tooltip = ToolTip(self.sidebar_clear_btn, "Tablo Temizle")
        self.toggle_tooltip = ToolTip(self.toggle_button, "Sidebar Aç / Kapat")
        self.refresh_tooltip = ToolTip(self.refresh_button, "Analizi Yenile")

        self.refresh_button.bind(
            "<Enter>",
            lambda e: self.refresh_button.configure(border_color="#93c5fd")
        )
        self.refresh_button.bind(
            "<Leave>",
            lambda e: self.refresh_button.configure(border_color=self.colors["border"])
        )

        # ================= FOOTER =================
        self.footer_frame = ctk.CTkFrame(
            self.main_frame,
            corner_radius=16,
            fg_color=self.colors["surface"]
        )
        self.footer_frame.grid(row=6, column=0, padx=16, pady=(0, 16), sticky="ew")
        self.footer_frame.grid_columnconfigure(0, weight=1)

        self.status_var = tk.StringVar(value="Uygulama başlatıldı.")
        self.status_label = ctk.CTkLabel(
            self.footer_frame,
            textvariable=self.status_var,
            text_color=self.colors["text_muted"],
            anchor="w"
        )
        self.status_label.grid(row=0, column=0, padx=16, pady=12, sticky="w")

        self.count_label = ctk.CTkLabel(
            self.footer_frame,
            text="0 kayıt",
            text_color=self.colors["text"]
        )
        self.count_label.grid(row=0, column=1, padx=16, pady=12, sticky="e")

        self.set_active_sidebar(None)
        self.set_sidebar_expanded_ui()

        buttons = [
            ("select", self.sidebar_select_btn, False),
            ("analyze", self.sidebar_analyze_btn, False),
            ("open", self.sidebar_open_btn, False),
            ("move", self.sidebar_move_btn, False),
            ("health", self.sidebar_health_btn, False),
            ("select_low", self.sidebar_select_low_btn, False),
            ("bulk_move", self.sidebar_bulk_move_btn, False),
            ("csv", self.sidebar_csv_btn, False),
            ("excel", self.sidebar_excel_btn, False),
            ("clear", self.sidebar_clear_btn, True),
        ]

        for name, btn, is_clear in buttons:
            self.bind_sidebar_button_hover(btn, name, is_clear=is_clear)

        self.update_empty_state()

    # ================= YARDIMCI =================
    def bind_card_hover(self, widget, normal_color, hover_color):
        widget.bind("<Enter>", lambda e: widget.configure(fg_color=hover_color))
        widget.bind("<Leave>", lambda e: widget.configure(fg_color=normal_color))

    def get_category(self, count):
        if 0 <= count <= 5:
            return "0-5"
        elif 6 <= count <= 25:
            return "6-25"
        elif 26 <= count <= 50:
            return "26-50"
        else:
            return "50+"

    def set_status(self, text, badge_text=None):
        self.status_var.set(text)

        if badge_text:
            self.status_badge.configure(text=badge_text)

            badge_map = {
                "Beklemede": ("#1e293b", "#e2e8f0"),
                "Seçildi": ("#0f3d73", "#dbeafe"),
                "Analiz": ("#92400e", "#fef3c7"),
                "Tamamlandı": ("#166534", "#dcfce7"),
                "Filtre": ("#4c1d95", "#ede9fe"),
                "Kaydedildi": ("#166534", "#dcfce7"),
                "Açıldı": ("#0f3d73", "#dbeafe"),
                "Taşındı": ("#166534", "#dcfce7"),
                "Uyarı": ("#92400e", "#fef3c7"),
                "Hata": ("#7f1d1d", "#fee2e2"),
            }

            fg, text_color = badge_map.get(badge_text, ("#1e293b", "#e2e8f0"))
            self.status_badge.configure(fg_color=fg, text_color=text_color)

    def apply_button_style(self, button, state="normal", is_clear=False):
        if is_clear:
            if state == "active":
                button.configure(
                    fg_color=self.colors["danger_hover"],
                    hover_color=self.colors["danger_hover"],
                    text_color="white",
                    border_width=1,
                    border_color="#fecaca"
                )
            elif state == "hover":
                button.configure(
                    fg_color="#7f1d1d",
                    hover_color="#7f1d1d",
                    text_color="white",
                    border_width=1,
                    border_color="#fca5a5"
                )
            else:
                button.configure(
                    fg_color="#3b1f24",
                    hover_color="#4a2329",
                    text_color=self.colors["text"],
                    border_width=1,
                    border_color="#6b2c35"
                )
            return

        if state == "active":
            button.configure(
                fg_color=self.colors["accent"],
                hover_color=self.colors["accent"],
                text_color="#111827",
                border_width=1,
                border_color="#fde68a"
            )
        elif state == "hover":
            button.configure(
                fg_color=self.colors["primary_hover"],
                hover_color=self.colors["primary_hover"],
                text_color="white",
                border_width=1,
                border_color="#93c5fd"
            )
        else:
            button.configure(
                fg_color=self.colors["surface_2"],
                hover_color=self.colors["surface_3"],
                text_color=self.colors["text"],
                border_width=1,
                border_color=self.colors["border"]
            )

    def bind_sidebar_button_hover(self, button, name, is_clear=False):
        def on_enter(event):
            if self.current_active == name:
                return
            self.apply_button_style(button, state="hover", is_clear=is_clear)

        def on_leave(event):
            if self.current_active == name:
                return
            self.apply_button_style(button, state="normal", is_clear=is_clear)

        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)

    def populate_tree(self, folder_details):
        for row in self.tree.get_children():
            self.tree.delete(row)

        sorted_folders = self.get_sorted_filtered_data(folder_details)
        self.current_filtered_data = sorted_folders

        for index, item in enumerate(sorted_folders):
            is_selected = item["path"] in self.selected_folders

            if is_selected:
                tag = "selected_low"
            else:
                tag = "evenrow" if index % 2 == 0 else "oddrow"

            self.tree.insert(
                "",
                "end",
                values=(item["folder"], item["count"], item["category"]),
                tags=(tag,)
            )

        self.tree.tag_configure("evenrow", background="#1e293b")
        self.tree.tag_configure("oddrow", background="#223046")
        self.tree.tag_configure(
            "selected_low",
            background="#7c2d12",
            foreground="white"
        )

        self.count_label.configure(text=f"{len(sorted_folders)} kayıt")
        self.update_tree_headings()
        self.update_empty_state()
        self.clear_preview_panel()

    def get_selected_tree_item(self):
        selected_item = self.tree.focus()

        if not selected_item:
            messagebox.showinfo("Bilgi", "Lütfen önce tablodan bir klasör seç.")
            return None

        values = self.tree.item(selected_item, "values")

        if not values:
            return None

        return values

    def clear_preview_panel(self, message="Tablodan bir klasör seç.\nBurada örnek görseller ve hızlı özet görünecek."):
        self.preview_info.configure(text=message)
        self.preview_footer.configure(text="İlk 6 görsel gösterilir")
        self.preview_thumbnails = []
        self.preview_image_names = []

        for widget in self.preview_images_frame.winfo_children():
            widget.destroy()

        self.preview_hint = ctk.CTkLabel(
            self.preview_images_frame,
            text="Önizleme için klasör seçilmedi",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=13, weight="bold")
        )
        self.preview_hint.grid(row=0, column=0, columnspan=3, padx=8, pady=16)

    def on_tree_select(self, event=None):
        selection = self.tree.selection()
        if not selection:
            return

        item_id = selection[0]
        values = self.tree.item(item_id, "values")
        if not values:
            return

        folder_name = values[0]
        folder_path = None

        for item in self.current_filtered_data:
            if item["folder"] == folder_name:
                folder_path = item["path"]
                break

        if folder_path:
            self.update_preview_panel(folder_path)

    def update_preview_panel(self, folder_path):
        try:
            files = os.listdir(folder_path)
        except Exception as error:
            self.clear_preview_panel(f"Önizleme alınamadı.\n{error}")
            return

        image_files = [f for f in files if f.lower().endswith((".jpg", ".jpeg", ".png"))]
        label_files = [f for f in files if f.lower().endswith(".txt")]

        image_names = {os.path.splitext(f)[0] for f in image_files}
        label_names = {os.path.splitext(f)[0] for f in label_files}
        empty_labels = 0

        for label_file in label_files:
            try:
                if os.path.getsize(os.path.join(folder_path, label_file)) == 0:
                    empty_labels += 1
            except Exception:
                pass

        if image_files and label_files and image_names == label_names and empty_labels == 0:
            health_text = "✅ Sağlıklı"
        elif not image_files:
            health_text = "⚠️ Görsel yok"
        elif image_names != label_names or empty_labels > 0:
            health_text = "⚠️ Sorunlu"
        else:
            health_text = "ℹ️ Kısmi"

        info_text = (
            f"Klasör: {os.path.basename(folder_path.rstrip('\\/'))}\n"
            f"Görüntü: {len(image_files)}\n"
            f"Etiket: {len(label_files)}\n"
            f"Durum: {health_text}"
        )
        self.preview_info.configure(text=info_text)
        self.preview_footer.configure(text=f"Konum: {folder_path}")

        for widget in self.preview_images_frame.winfo_children():
            widget.destroy()

        if not image_files:
            empty_label = ctk.CTkLabel(
                self.preview_images_frame,
                text="Bu klasörde önizlenecek görsel yok",
                text_color=self.colors["text_muted"],
                font=ctk.CTkFont(size=13, weight="bold")
            )
            empty_label.grid(row=0, column=0, columnspan=3, padx=8, pady=16)
            self.preview_thumbnails = []
            self.preview_image_names = []
            return

        preview_files = sorted(image_files)[:6]
        self.preview_thumbnails = []
        self.preview_image_names = preview_files

        for index, image_name in enumerate(preview_files):
            image_path = os.path.join(folder_path, image_name)
            try:
                image = Image.open(image_path)

                from PIL import ImageDraw

                img = Image.open(image_path)

                # label var mı?
                label_path = os.path.splitext(image_path)[0] + ".txt"

                if os.path.exists(label_path):
                    draw = ImageDraw.Draw(img)

                    with open(label_path, "r") as f:
                        lines = f.readlines()

                    for line in lines:
                        parts = line.strip().split()
                        if len(parts) != 5:
                            continue

                        _, x, y, w, h = map(float, parts)

                        img_w, img_h = img.size

                        x_center = x * img_w
                        y_center = y * img_h
                        box_w = w * img_w
                        box_h = h * img_h

                        x1 = x_center - box_w / 2
                        y1 = y_center - box_h / 2
                        x2 = x_center + box_w / 2
                        y2 = y_center + box_h / 2

                        draw.rectangle([x1, y1, x2, y2], outline="red", width=2)

                image.thumbnail((120, 120))
                ctk_image = ctk.CTkImage(light_image=image, dark_image=image, size=(110, 110))
                self.preview_thumbnails.append(ctk_image)

                cell = ctk.CTkFrame(
                    self.preview_images_frame,
                    corner_radius=10,
                    fg_color=self.colors["surface"],
                    border_width=1,
                    border_color=self.colors["border"]
                )
                cell.grid(row=index // 3, column=index % 3, padx=6, pady=6, sticky="nsew")

                img_label = ctk.CTkLabel(cell, image=ctk_image, text="")
                img_label.pack(padx=6, pady=(6, 4))

                name_label = ctk.CTkLabel(
                    cell,
                    text=image_name[:18] + ("..." if len(image_name) > 18 else ""),
                    text_color=self.colors["text_muted"],
                    font=ctk.CTkFont(size=10)
                )
                name_label.pack(padx=4, pady=(0, 6))
            except Exception:
                error_cell = ctk.CTkLabel(
                    self.preview_images_frame,
                    text=f"{image_name}\naçılamadı",
                    text_color="#fca5a5",
                    font=ctk.CTkFont(size=10, weight="bold"),
                    justify="center"
                )
                error_cell.grid(row=index // 3, column=index % 3, padx=6, pady=6)

    def refresh_analysis(self):
        if self.is_busy:
            return

        if not self.selected_folder:
            messagebox.showinfo("Bilgi", "Önce bir klasör seçmelisin.")
            self.set_status("Yenileme başlatılamadı.", "Uyarı")
            return

        self.run_analysis_async(is_refresh=True)

    # ================= TEMEL İŞLEMLER =================
    def select_folder(self):
        self.flash_sidebar_action("select")
        selected = filedialog.askdirectory()

        if selected:
            self.selected_folder = selected
            self.folder_info_var.set(selected)
            self.set_status("Klasör seçildi.", "Seçildi")
        else:
            self.set_status("Klasör seçme işlemi iptal edildi.", "Beklemede")

    def analyze_selected_folder(self):
        if self.is_busy:
            return

        self.flash_sidebar_action("analyze")

        if not self.selected_folder:
            messagebox.showwarning("Uyarı", "Önce bir klasör seçmelisin.")
            self.set_status("Analiz başlatılamadı.", "Uyarı")
            return

        self.run_analysis_async(is_refresh=False)

    def run_analysis_async(self, is_refresh=False):
        action_text = "Analiz yenileniyor..." if is_refresh else "Analiz başlatıldı..."
        self.set_busy_state(True, action_text)

        def safe_progress(progress_ratio, current, total, message):
            self.root.after(0, lambda: self.update_analysis_progress(progress_ratio, current, total, message))

        def safe_status(message):
            self.root.after(0, lambda: self.loading_message_var.set(message))

        def worker():
            try:
                mode = self.analysis_mode.get()
                analyzer = DataFolderAnalyzer(
                    self.selected_folder,
                    mode=mode,
                    progress_callback=safe_progress,
                    status_callback=safe_status
                )
                result = analyzer.analyze()
                self.root.after(0, lambda: self.on_analysis_success(result, is_refresh))
            except Exception as error:
                self.root.after(0, lambda: self.on_analysis_error(error, is_refresh))

        threading.Thread(target=worker, daemon=True).start()

    def update_analysis_progress(self, progress_ratio, current, total, message):
        progress_ratio = max(0.0, min(progress_ratio, 1.0))
        self.set_progress_value(progress_ratio)
        self.loading_message_var.set(message)
        self.loading_detail_var.set(f"{current} / {total} klasör işlendi")

    def on_analysis_success(self, result, is_refresh=False):
        self.show_result(result)
        done_text = "Analiz yenilendi." if is_refresh else "Analiz tamamlandı."
        self.set_status(done_text, "Tamamlandı")
        self.loading_message_var.set("Analiz tamamlandı")
        self.loading_detail_var.set("Sonuçlar tabloya aktarılıyor")
        self.root.after(180, lambda: self.set_busy_state(False))

    def on_analysis_error(self, error, is_refresh=False):
        self.set_busy_state(False)
        action_name = "Yenileme" if is_refresh else "Analiz"
        messagebox.showerror("Hata", f"{action_name} sırasında hata oluştu:\n{error}")
        self.set_status(f"{action_name} başarısız.", "Hata")

    def show_result(self, result):
        self.last_result = result
        self.selected_folders.clear()
        self.sort_column = "count"
        self.sort_reverse = True
        categories, folder_details, total_images, total_folders = result

        self.animate_counter(self.total_folders_value, total_folders)
        self.animate_counter(self.total_images_value, total_images)
        self.animate_counter(self.low_count_value, categories["0-5"])
        self.animate_counter(self.high_count_value, categories["50+"])

        self.search_var.set("")
        self.filter_segmented.set("Tümü")

        self.populate_tree(folder_details)
        self.clear_preview_panel()

        self.summary_var.set(
            f"{total_folders} klasör | {total_images} görüntü | "
            f"0-5: {categories['0-5']} | "
            f"6-25: {categories['6-25']} | "
            f"26-50: {categories['26-50']} | "
            f"50+: {categories['50+']}"
        )

        self.update_empty_state()

    def animate_counter(self, label_widget, target_value, duration=350):
        start_value = 0
        try:
            start_value = int(label_widget.cget("text"))
        except Exception:
            start_value = 0

        steps = max(1, min(16, duration // 20))
        delta = target_value - start_value

        def step(index=0):
            if index >= steps:
                label_widget.configure(text=str(target_value))
                return

            current_value = start_value + int(delta * ((index + 1) / steps))
            label_widget.configure(text=str(current_value))
            self.root.after(20, lambda: step(index + 1))

        step()

    def clear_list(self):
        self.flash_sidebar_action("clear")

        for row in self.tree.get_children():
            self.tree.delete(row)

        self.last_result = None
        self.selected_folders.clear()
        self.current_filtered_data = []
        self.search_var.set("")
        self.sort_column = None
        self.sort_reverse = False
        self.summary_var.set("Durum: Tablo temizlendi")
        self.count_label.configure(text="0 kayıt")
        self.filter_segmented.set("Tümü")
        self.set_status("Tablo temizlendi.", "Beklemede")

        self.total_folders_value.configure(text="0")
        self.total_images_value.configure(text="0")
        self.low_count_value.configure(text="0")
        self.high_count_value.configure(text="0")

        self.update_tree_headings()
        self.update_empty_state()
        self.clear_preview_panel()

    # ================= FİLTRE =================
    def on_filter_change(self, value):
        self.apply_filters()

    # ================= KLASÖR AÇ =================
    def open_selected_folder(self, event=None):
        self.flash_sidebar_action("open")

        if not self.selected_folder:
            messagebox.showinfo("Bilgi", "Önce analiz yapılmış bir klasör seçmelisin.")
            return

        selected_values = self.get_selected_tree_item()
        if not selected_values:
            return

        folder_name = selected_values[0]
        folder_path = os.path.join(self.selected_folder, folder_name)

        if not os.path.exists(folder_path):
            messagebox.showerror("Hata", f"Klasör bulunamadı:\n{folder_path}")
            return

        try:
            os.startfile(folder_path)
            self.set_status(f"Klasör açıldı: {folder_name}", "Açıldı")
        except Exception as error:
            messagebox.showerror("Hata", f"Klasör açılamadı:\n{error}")

    # ================= CSV =================
    def export_to_csv(self):
        self.flash_sidebar_action("csv")

        if not self.last_result:
            messagebox.showinfo("Bilgi", "Önce analiz yapmalısın.")
            return

        categories, folder_details, total_images, total_folders = self.last_result
        # filtrelenmiş veri varsa onu kullan
        if self.current_filtered_data:
            folder_details = self.current_filtered_data
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Dosyası", "*.csv")],
            title="Analizi CSV olarak kaydet (.csv)"
        )

        if not file_path:
            self.set_status("CSV kaydetme iptal edildi.", "Beklemede")
            return

        try:
            sorted_folders = sorted(folder_details, key=lambda x: x["count"], reverse=True)

            with open(file_path, mode="w", newline="", encoding="utf-8-sig") as file:
                writer = csv.writer(file, delimiter=";")

                writer.writerow(["Sıra No", "Klasör", "Görüntü Sayısı", "Kategori", "Tam Yol"])

                for index, item in enumerate(sorted_folders, start=1):
                    writer.writerow([
                        index,
                        item["folder"],
                        item["count"],
                        item["category"],
                        item["path"]
                    ])

            self.set_status(f"CSV kaydedildi: {os.path.basename(file_path)}", "Kaydedildi")
            messagebox.showinfo("Başarılı", f"CSV başarıyla kaydedildi:\n{file_path}")

        except Exception as error:
            messagebox.showerror("Hata", f"CSV kaydedilemedi:\n{error}")
            self.set_status("CSV dışa aktarma başarısız.", "Hata")

    # ================= EXCEL =================
    def export_to_excel(self):
        self.flash_sidebar_action("excel")

        if not self.last_result:
            messagebox.showinfo("Bilgi", "Önce analiz yapmalısın.")
            return

        categories, folder_details, total_images, total_folders = self.last_result
        # filtrelenmiş veri varsa onu kullan
        if self.current_filtered_data:
            folder_details = self.current_filtered_data
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx")],
            title="Analizi Excel olarak kaydet (.xlsx)"
        )

        if not file_path:
            self.set_status("Excel kaydetme iptal edildi.", "Beklemede")
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Analiz Sonucu"

            header_fill = PatternFill(fill_type="solid", start_color="2D2D2D", end_color="2D2D2D")
            header_font = Font(bold=True, color="FFFFFF")
            title_font = Font(size=14, bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")

            ws["A1"] = "Klasör Analiz Raporu"
            ws["A1"].font = title_font

            ws["A3"] = "Seçilen Klasör"
            ws["B3"] = self.selected_folder

            ws["A4"] = "Toplam Klasör"
            ws["B4"] = total_folders

            ws["A5"] = "Toplam Görüntü"
            ws["B5"] = total_images

            ws["D3"] = "Kategori"
            ws["E3"] = "Klasör Sayısı"

            for cell in ["D3", "E3"]:
                ws[cell].fill = header_fill
                ws[cell].font = header_font
                ws[cell].alignment = center_align

            ws["D4"] = "0-5"
            ws["E4"] = categories["0-5"]
            ws["D5"] = "6-25"
            ws["E5"] = categories["6-25"]
            ws["D6"] = "26-50"
            ws["E6"] = categories["26-50"]
            ws["D7"] = "50+"
            ws["E7"] = categories["50+"]

            start_row = 10
            headers = ["Sıra No", "Klasör", "Görüntü Sayısı", "Kategori", "Tam Yol"]

            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            sorted_folders = sorted(folder_details, key=lambda x: x["count"], reverse=True)

            for index, item in enumerate(sorted_folders, start=1):
                row = start_row + index
                ws.cell(row=row, column=1, value=index)
                ws.cell(row=row, column=2, value=item["folder"])
                ws.cell(row=row, column=3, value=item["count"])
                ws.cell(row=row, column=4, value=item["category"])
                ws.cell(row=row, column=5, value=item["path"])

                ws.cell(row=row, column=1).alignment = center_align
                ws.cell(row=row, column=3).alignment = center_align
                ws.cell(row=row, column=4).alignment = center_align
                ws.cell(row=row, column=2).alignment = left_align
                ws.cell(row=row, column=5).alignment = left_align

            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 35
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 70

            last_row = start_row + len(sorted_folders)
            ws.auto_filter.ref = f"A{start_row}:E{last_row}"

            ws.row_dimensions[1].height = 24
            ws.row_dimensions[start_row].height = 22

            wb.save(file_path)

            self.set_status(f"Excel kaydedildi: {os.path.basename(file_path)}", "Kaydedildi")
            messagebox.showinfo("Başarılı", f"Excel başarıyla kaydedildi:\n{file_path}")

        except Exception as error:
            messagebox.showerror("Hata", f"Excel kaydedilemedi:\n{error}")
            self.set_status("Excel dışa aktarma başarısız.", "Hata")

    def run_health_check(self):
        self.flash_sidebar_action("health")

        if not self.last_result:
            messagebox.showinfo("Bilgi", "Önce analiz yapmalısın.")
            return

        _, folder_details, _, _ = self.last_result

        total_images = 0
        total_labels = 0
        missing_labels = 0
        missing_images = 0
        empty_labels = 0

        self.selected_folders.clear()
        self.valid_folders.clear()
        detailed_report_lines = []

        # Ana klasörü de dahil et + alt klasörleri ekle
        paths_to_check = []
        seen_paths = set()

        if self.selected_folder and os.path.isdir(self.selected_folder):
            paths_to_check.append(self.selected_folder)
            seen_paths.add(os.path.normpath(self.selected_folder))

        for item in folder_details:
            folder_path = item["path"]
            normalized = os.path.normpath(folder_path)

            if normalized not in seen_paths and os.path.isdir(folder_path):
                paths_to_check.append(folder_path)
                seen_paths.add(normalized)

        for folder_path in paths_to_check:
            folder_name = os.path.basename(folder_path.rstrip("\\/")) or folder_path

            try:
                files = os.listdir(folder_path)
            except Exception:
                continue

            images = [f for f in files if f.lower().endswith((".jpg", ".jpeg", ".png"))]
            labels = [f for f in files if f.lower().endswith(".txt")]

            total_images += len(images)
            total_labels += len(labels)

            image_map = {os.path.splitext(f)[0]: f for f in images}
            label_map = {os.path.splitext(f)[0]: f for f in labels}

            image_names = set(image_map.keys())
            label_names = set(label_map.keys())

            folder_problem_lines = []

            # Etiketi eksik görüntüler
            missing_label_names = sorted(image_names - label_names)
            if missing_label_names:
                missing_labels += len(missing_label_names)
                folder_problem_lines.append("- Etiketi eksik görüntüler:")
                for base_name in missing_label_names:
                    image_file = image_map[base_name]
                    folder_problem_lines.append(
                        f"  • {image_file} -> beklenen label: {base_name}.txt"
                    )

            # Görüntüsü eksik etiketler
            missing_image_names = sorted(label_names - image_names)
            if missing_image_names:
                missing_images += len(missing_image_names)
                folder_problem_lines.append("- Görüntüsü eksik etiketler:")
                for base_name in missing_image_names:
                    label_file = label_map[base_name]
                    folder_problem_lines.append(
                        f"  • {label_file} -> beklenen image: {base_name}.jpg / .jpeg / .png"
                    )

            # Boş etiket dosyaları
            empty_label_files = []
            for label_file in labels:
                try:
                    full_label_path = os.path.join(folder_path, label_file)
                    if os.path.getsize(full_label_path) == 0:
                        empty_labels += 1
                        empty_label_files.append(label_file)
                except Exception:
                    pass

            if empty_label_files:
                folder_problem_lines.append("- Boş etiket dosyaları:")
                for label_file in empty_label_files:
                    folder_problem_lines.append(f"  • {label_file}")

            # Sağlıklı klasör: image var, txt var, isimler birebir eşleşiyor, boş txt yok
            if images and labels and (image_names == label_names) and not empty_label_files:
                self.valid_folders.add(folder_path)

            # Problemli klasör
            if folder_problem_lines:
                self.selected_folders.add(folder_path)
                detailed_report_lines.append(f"Klasör: {folder_name}")
                detailed_report_lines.extend(folder_problem_lines)
                detailed_report_lines.append("")

        self.apply_filters()

        summary_text = (
            f"Toplam Görüntü (Image): {total_images}\n"
            f"Toplam Etiket (Label): {total_labels}\n\n"
            f"Etiketi eksik görüntü sayısı: {missing_labels}\n"
            f"Görüntüsü eksik etiket sayısı: {missing_images}\n"
            f"Boş etiket dosyası sayısı: {empty_labels}\n"
            f"Problemli klasör sayısı: {len(self.selected_folders)}\n"
            f"Sağlıklı klasör sayısı: {len(self.valid_folders)}"
        )

        if detailed_report_lines:
            details_text = "\n".join(detailed_report_lines)
        else:
            details_text = "Sorun bulunmadı 🎉"

        self.show_health_report_window(summary_text, details_text)

    def show_health_report_window(self, summary_text, details_text):
        window = ctk.CTkToplevel(self.root)
        window.title("Dataset Sağlık Raporu")
        window.geometry("900x650")
        window.minsize(760, 520)
        window.grab_set()

        window.grid_rowconfigure(1, weight=1)
        window.grid_columnconfigure(0, weight=1)

        header_frame = ctk.CTkFrame(
            window,
            corner_radius=14,
            fg_color=self.colors["surface"]
        )
        header_frame.grid(row=0, column=0, padx=16, pady=(16, 10), sticky="ew")
        header_frame.grid_columnconfigure(0, weight=1)

        title_label = ctk.CTkLabel(
            header_frame,
            text="🩺 Dataset Sağlık Raporu",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.grid(row=0, column=0, padx=16, pady=(14, 6), sticky="w")

        summary_label = ctk.CTkLabel(
            header_frame,
            text=summary_text,
            text_color=self.colors["text_muted"],
            justify="left",
            anchor="w",
            font=ctk.CTkFont(size=13)
        )
        summary_label.grid(row=1, column=0, padx=16, pady=(0, 14), sticky="ew")

        content_frame = ctk.CTkFrame(
            window,
            corner_radius=14,
            fg_color=self.colors["surface"]
        )
        content_frame.grid(row=1, column=0, padx=16, pady=(0, 10), sticky="nsew")
        content_frame.grid_rowconfigure(0, weight=1)
        content_frame.grid_columnconfigure(0, weight=1)

        text_box = ctk.CTkTextbox(
            content_frame,
            wrap="word",
            fg_color=self.colors["surface_2"],
            text_color=self.colors["text"],
            corner_radius=12,
            border_width=1,
            border_color=self.colors["border"],
            font=ctk.CTkFont(size=13)
        )
        text_box.grid(row=0, column=0, padx=14, pady=14, sticky="nsew")

        text_box.insert("1.0", details_text)
        text_box.configure(state="disabled")

        footer_frame = ctk.CTkFrame(
            window,
            corner_radius=14,
            fg_color=self.colors["surface"]
        )
        footer_frame.grid(row=2, column=0, padx=16, pady=(0, 16), sticky="ew")
        footer_frame.grid_columnconfigure(0, weight=1)

        copy_button = ctk.CTkButton(
            footer_frame,
            text="📋 Raporu Kopyala",
            command=lambda: self.copy_health_report(summary_text, details_text),
            corner_radius=10,
            fg_color=self.colors["primary"],
            hover_color=self.colors["primary_hover"],
            text_color="white"
        )
        copy_button.grid(row=0, column=0, padx=(16, 8), pady=12, sticky="w")

        show_labeled_button = ctk.CTkButton(
            footer_frame,
            text="✅ Etiketlileri Göster",
            command=self.show_labeled_folders,
            corner_radius=10,
            fg_color=self.colors["success"],
            hover_color="#15803d",
            text_color="white"
        )
        show_labeled_button.grid(row=0, column=1, padx=(8, 8), pady=12, sticky="w")

        close_button = ctk.CTkButton(
            footer_frame,
            text="Kapat",
            command=window.destroy,
            corner_radius=10,
            fg_color=self.colors["surface_2"],
            hover_color=self.colors["surface_3"],
            text_color=self.colors["text"],
            border_width=1,
            border_color=self.colors["border"]
        )
        close_button.grid(row=0, column=2, padx=(8, 16), pady=12, sticky="e")

    def show_labeled_folders_window(self, labeled_folders):
        window = ctk.CTkToplevel(self.root)
        window.title("Etiketli / Sağlıklı Klasörler")
        window.geometry("900x600")
        window.minsize(760, 500)
        window.grab_set()

        window.grid_rowconfigure(1, weight=1)
        window.grid_columnconfigure(0, weight=1)

        header_frame = ctk.CTkFrame(
            window,
            corner_radius=14,
            fg_color=self.colors["surface"]
        )
        header_frame.grid(row=0, column=0, padx=16, pady=(16, 10), sticky="ew")
        header_frame.grid_columnconfigure(0, weight=1)

        title_label = ctk.CTkLabel(
            header_frame,
            text="✅ Etiketli / Sağlıklı Klasörler",
            text_color=self.colors["text"],
            font=ctk.CTkFont(size=22, weight="bold")
        )
        title_label.grid(row=0, column=0, padx=16, pady=(14, 4), sticky="w")

        info_label = ctk.CTkLabel(
            header_frame,
            text=f"Toplam {len(labeled_folders)} klasör bulundu.",
            text_color=self.colors["text_muted"],
            font=ctk.CTkFont(size=13)
        )
        info_label.grid(row=1, column=0, padx=16, pady=(0, 14), sticky="w")

        table_frame = ctk.CTkFrame(
            window,
            corner_radius=14,
            fg_color=self.colors["surface"]
        )
        table_frame.grid(row=1, column=0, padx=16, pady=(0, 10), sticky="nsew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        tree = ttk.Treeview(
            table_frame,
            columns=("folder", "count", "category", "path"),
            show="headings"
        )

        tree.heading("folder", text="Klasör")
        tree.heading("count", text="Görüntü Sayısı")
        tree.heading("category", text="Kategori")
        tree.heading("path", text="Tam Yol")

        tree.column("folder", width=220, anchor="w")
        tree.column("count", width=120, anchor="center")
        tree.column("category", width=100, anchor="center")
        tree.column("path", width=420, anchor="w")

        tree.grid(row=0, column=0, sticky="nsew")

        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=scrollbar.set)

        for index, item in enumerate(labeled_folders):
            tag = "evenrow" if index % 2 == 0 else "oddrow"
            tree.insert(
                "",
                "end",
                values=(item["folder"], item["count"], item["category"], item["path"]),
                tags=(tag,)
            )

        tree.tag_configure("evenrow", background="#1e293b", foreground="white")
        tree.tag_configure("oddrow", background="#223046", foreground="white")

        footer_frame = ctk.CTkFrame(
            window,
            corner_radius=14,
            fg_color=self.colors["surface"]
        )
        footer_frame.grid(row=2, column=0, padx=16, pady=(0, 16), sticky="ew")
        footer_frame.grid_columnconfigure(0, weight=1)

        def copy_labeled_list():
            lines = []
            for item in labeled_folders:
                lines.append(
                    f"Klasör: {item['folder']} | Görüntü: {item['count']} | "
                    f"Kategori: {item['category']} | Yol: {item['path']}"
                )

            text = "\n".join(lines) if lines else "Kayıt bulunamadı."
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            self.root.update()
            messagebox.showinfo("Bilgi", "Etiketli klasör listesi panoya kopyalandı.")

        copy_button = ctk.CTkButton(
            footer_frame,
            text="📋 Listeyi Kopyala",
            command=copy_labeled_list,
            corner_radius=10,
            fg_color=self.colors["primary"],
            hover_color=self.colors["primary_hover"],
            text_color="white"
        )
        copy_button.grid(row=0, column=0, padx=(16, 8), pady=12, sticky="w")

        close_button = ctk.CTkButton(
            footer_frame,
            text="Kapat",
            command=window.destroy,
            corner_radius=10,
            fg_color=self.colors["surface_2"],
            hover_color=self.colors["surface_3"],
            text_color=self.colors["text"],
            border_width=1,
            border_color=self.colors["border"]
        )
        close_button.grid(row=0, column=1, padx=(8, 16), pady=12, sticky="e")

    def copy_health_report(self, summary_text, details_text):
        full_text = f"{summary_text}\n\nDETAYLI PROBLEM LİSTESİ\n\n{details_text}"
        self.root.clipboard_clear()
        self.root.clipboard_append(full_text)
        self.root.update()
        messagebox.showinfo("Bilgi", "Sağlık raporu panoya kopyalandı.")

    def show_labeled_folders(self):
        if not self.valid_folders:
            messagebox.showinfo("Bilgi", "Sağlıklı / tam eşleşen etiketli klasör bulunamadı.")
            return

        if not self.last_result:
            messagebox.showinfo("Bilgi", "Önce analiz yapmalısın.")
            return

        _, folder_details, _, _ = self.last_result

        labeled_only = []

        # Ana klasör de sağlık kontrolünde valid olabilir ama folder_details içinde olmayabilir
        if self.selected_folder in self.valid_folders:
            try:
                files = os.listdir(self.selected_folder)
                images = [f for f in files if f.lower().endswith((".jpg", ".jpeg", ".png"))]
                category = self.get_category(len(images))

                labeled_only.append({
                    "folder": os.path.basename(self.selected_folder.rstrip("\\/")) or self.selected_folder,
                    "count": len(images),
                    "category": category,
                    "path": self.selected_folder
                })
            except Exception:
                pass

        # Alt klasörlerden sağlıklı olanları ekle
        for item in folder_details:
            if item["path"] in self.valid_folders:
                labeled_only.append(item)

        # Aynı path iki kez eklenmesin
        unique_labeled = []
        seen_paths = set()

        for item in labeled_only:
            normalized = os.path.normpath(item["path"])
            if normalized not in seen_paths:
                unique_labeled.append(item)
                seen_paths.add(normalized)

        self.show_labeled_folders_window(unique_labeled)

    # ================= TAŞIMA =================
    def select_low_folders(self):
        self.flash_sidebar_action("select_low")

        if not self.last_result:
            messagebox.showinfo("Bilgi", "Önce analiz yapmalısın.")
            return

        try:
            threshold = int(self.low_threshold_var.get().strip())
            if threshold < 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Uyarı", "Threshold için geçerli bir sayı gir.")
            return

        categories, folder_details, total_images, total_folders = self.last_result

        self.selected_folders.clear()

        for item in folder_details:
            if item["count"] <= threshold:
                self.selected_folders.add(item["path"])

        self.apply_filters()
        self.set_status(f"{len(self.selected_folders)} klasör seçildi (≤ {threshold}).", "Filtre")
        messagebox.showinfo(
            "Bilgi",
            f"{len(self.selected_folders)} adet klasör seçildi.\nThreshold: {threshold}"
        )

    def move_selected_folder(self):
        self.flash_sidebar_action("move")

        if not self.selected_folder:
            messagebox.showinfo("Bilgi", "Önce analiz yapılmış bir klasör seçmelisin.")
            return

        selected_values = self.get_selected_tree_item()
        if not selected_values:
            return

        folder_name = selected_values[0]
        source_path = os.path.join(self.selected_folder, folder_name)

        if not os.path.exists(source_path):
            messagebox.showerror("Hata", f"Kaynak klasör bulunamadı:\n{source_path}")
            return

        target_directory = filedialog.askdirectory(title="Hedef klasörü seç")
        if not target_directory:
            self.set_status("Taşıma işlemi iptal edildi.", "Beklemede")
            return

        destination_path = os.path.join(target_directory, folder_name)

        if os.path.exists(destination_path):
            messagebox.showwarning(
                "Uyarı",
                f"Hedefte aynı isimde klasör zaten var:\n{destination_path}"
            )
            return

        confirm = messagebox.askyesno(
            "Onay",
            f"'{folder_name}' klasörü taşınsın mı?\n\n"
            f"Kaynak:\n{source_path}\n\n"
            f"Hedef:\n{destination_path}"
        )

        if not confirm:
            self.set_status("Taşıma işlemi iptal edildi.", "Beklemede")
            return

        try:
            shutil.move(source_path, destination_path)
            self.set_status(f"Klasör taşındı: {folder_name}", "Taşındı")
            self.refresh_analysis()
            messagebox.showinfo("Başarılı", f"Klasör taşındı:\n{folder_name}")

        except Exception as error:
            messagebox.showerror("Hata", f"Klasör taşınamadı:\n{error}")
            self.set_status("Taşıma işlemi başarısız.", "Hata")

    def move_selected_folders_bulk(self):
        self.flash_sidebar_action("bulk_move")

        if not self.selected_folders:
            messagebox.showwarning("Uyarı", "Önce 'Az Verili Seç' ile klasör seçmelisin.")
            return

        target_directory = filedialog.askdirectory(title="Hedef klasörü seç")
        if not target_directory:
            self.set_status("Toplu taşıma iptal edildi.", "Beklemede")
            return

        moved = 0
        failed = []

        for folder_path in list(self.selected_folders):
            try:
                folder_name = os.path.basename(folder_path.rstrip("\\/"))
                destination = os.path.join(target_directory, folder_name)

                if os.path.exists(destination):
                    failed.append(f"{folder_name} (hedefte zaten var)")
                    continue

                shutil.move(folder_path, destination)
                moved += 1

            except Exception as error:
                failed.append(f"{folder_name} ({error})")

        self.selected_folders.clear()

        message = f"{moved} klasör taşındı."
        if failed:
            message += f"\n{len(failed)} klasör taşınamadı."

        messagebox.showinfo("Sonuç", message)
        self.set_status(f"{moved} klasör toplu taşındı.", "Taşındı")
        self.refresh_analysis()

    # ================= SIDEBAR STYLE =================
    def set_active_sidebar(self, active_name=None):
        self.current_active = active_name

        button_map = {
            "select": self.sidebar_select_btn,
            "analyze": self.sidebar_analyze_btn,
            "open": self.sidebar_open_btn,
            "move": self.sidebar_move_btn,
            "health": self.sidebar_health_btn,
            "select_low": self.sidebar_select_low_btn,
            "bulk_move": self.sidebar_bulk_move_btn,
            "csv": self.sidebar_csv_btn,
            "excel": self.sidebar_excel_btn,
            "clear": self.sidebar_clear_btn,
        }

        for name, button in button_map.items():
            is_clear = (name == "clear")
            if name == active_name:
                self.apply_button_style(button, state="active", is_clear=is_clear)
            else:
                self.apply_button_style(button, state="normal", is_clear=is_clear)

            button.grid_configure(padx=14, pady=6)

    def reset_sidebar_highlight(self):
        self.set_active_sidebar(None)

    def toggle_sidebar(self):
        if self.sidebar_expanded:
            self.set_sidebar_collapsed_ui()
            self.sidebar_frame.configure(width=self.sidebar_collapsed_width)
            self.root.grid_columnconfigure(0, minsize=self.sidebar_collapsed_width)
            self.toggle_button.configure(text="⮞")
            self.sidebar_expanded = False
        else:
            self.set_sidebar_expanded_ui()
            self.sidebar_frame.configure(width=self.sidebar_expanded_width)
            self.root.grid_columnconfigure(0, minsize=self.sidebar_expanded_width)
            self.toggle_button.configure(text="☰")
            self.sidebar_expanded = True

        self.root.update_idletasks()

    def set_sidebar_collapsed_ui(self):
        self.logo_label.configure(text="🧠")
        self.sidebar_subtitle.configure(text="")
        self.nav_title.configure(text="")
        self.export_title.configure(text="")
        self.actions_title.configure(text="")
        self.sidebar_info.configure(text="")

        self.sidebar_select_btn.configure(text="📂", anchor="center")
        self.sidebar_analyze_btn.configure(text="🔍", anchor="center")
        self.sidebar_open_btn.configure(text="📁", anchor="center")
        self.sidebar_move_btn.configure(text="🚚", anchor="center")
        self.sidebar_health_btn.configure(text="🩺", anchor="center")
        self.sidebar_select_low_btn.configure(text="⚡", anchor="center")
        self.sidebar_bulk_move_btn.configure(text="🚀", anchor="center")
        self.sidebar_csv_btn.configure(text="🧾", anchor="center")
        self.sidebar_excel_btn.configure(text="📊", anchor="center")
        self.sidebar_clear_btn.configure(text="🧹", anchor="center")

    def set_sidebar_expanded_ui(self):
        self.logo_label.configure(text="🧠 Analyzer")
        self.sidebar_subtitle.configure(text="Klasör analiz ve yönetim paneli")
        self.nav_title.configure(text="MENÜ")
        self.export_title.configure(text="DIŞA AKTAR")
        self.actions_title.configure(text="İŞLEMLER")
        self.sidebar_info.configure(text="Beklemede\nBir klasör seçip analize başlayabilirsin.")

        self.sidebar_select_btn.configure(text="📂 Klasör Seç", anchor="w")
        self.sidebar_analyze_btn.configure(text="🔍 Analiz Et", anchor="w")
        self.sidebar_open_btn.configure(text="📁 Seçiliyi Aç", anchor="w")
        self.sidebar_move_btn.configure(text="🚚 Klasörü Taşı", anchor="w")
        self.sidebar_health_btn.configure(text="🩺 Sağlık Kontrolü", anchor="w")
        self.sidebar_select_low_btn.configure(text="⚡ Az Verili Seç", anchor="w")
        self.sidebar_bulk_move_btn.configure(text="🚀 Seçilenleri Taşı", anchor="w")
        self.sidebar_csv_btn.configure(text="🧾 CSV Kaydet", anchor="w")
        self.sidebar_excel_btn.configure(text="📊 Excel Kaydet", anchor="w")
        self.sidebar_clear_btn.configure(text="🧹 Tablo Temizle", anchor="w")

    def flash_sidebar_action(self, name, duration=220):
        self.set_active_sidebar(name)
        self.root.after(duration, self.reset_sidebar_highlight)

    def get_action_buttons(self):
        return [
            self.sidebar_select_btn,
            self.sidebar_analyze_btn,
            self.sidebar_open_btn,
            self.sidebar_move_btn,
            self.sidebar_health_btn,
            self.sidebar_select_low_btn,
            self.sidebar_bulk_move_btn,
            self.sidebar_csv_btn,
            self.sidebar_excel_btn,
            self.sidebar_clear_btn,
            self.refresh_button,
        ]

    def set_busy_state(self, busy=True, message="İşlem devam ediyor..."):
        self.is_busy = busy

        if busy:
            self.set_status(message, "Analiz")
            self.status_badge.configure(text="Yükleniyor...")
            self.root.configure(cursor="watch")
            self.start_progress(indeterminate=False)
            self.show_loading_overlay()
        else:
            self.root.configure(cursor="")
            self.stop_progress()
            self.hide_loading_overlay()

            if self.status_badge.cget("text") == "Yükleniyor...":
                self.status_badge.configure(text="Tamamlandı")

        for btn in self.get_action_buttons():
            btn.configure(state="disabled" if busy else "normal")

        if not busy:
            self.set_active_sidebar(None)

    def show_loading_overlay(self):
        self.loading_message_var.set("Klasörler hazırlanıyor...")
        self.loading_detail_var.set("0 / 0 klasör işlendi")
        self.loading_overlay.lift()
        self.loading_overlay.grid()
        self.animate_loading_spinner()
        self.animate_loading_text()

    def hide_loading_overlay(self):
        if self._loading_frame_job:
            self.root.after_cancel(self._loading_frame_job)
            self._loading_frame_job = None

        if self._loading_text_job:
            self.root.after_cancel(self._loading_text_job)
            self._loading_text_job = None

        self.loading_spinner_label.configure(text="◐")
        self.loading_overlay.grid_remove()

    def animate_loading_spinner(self):
        frames = ["◐", "◓", "◑", "◒"]
        self.loading_spinner_label.configure(text=frames[self._loading_spinner_index % len(frames)])
        self._loading_spinner_index += 1

        if self.is_busy:
            self._loading_frame_job = self.root.after(120, self.animate_loading_spinner)

    def animate_loading_text(self):
        dots = "." * ((self._loading_spinner_index % 3) + 1)
        current = self.loading_message_var.get().rstrip(".")

        if current:
            self.loading_title_label.configure(text="Analiz sürüyor")

        if self.is_busy and current and "Taranıyor" not in current and "tamamlandı" not in current.lower():
            self.loading_message_label.configure(text=current + dots)
        else:
            self.loading_message_label.configure(text=self.loading_message_var.get())

        if self.is_busy:
            self._loading_text_job = self.root.after(280, self.animate_loading_text)

    def update_empty_state(self):
        has_rows = len(self.tree.get_children()) > 0

        if has_rows:
            self.empty_placeholder.grid_remove()
        else:
            self.empty_placeholder.grid()

    def apply_filters(self):
        if not self.last_result:
            self.populate_tree([])
            return

        categories, folder_details, total_images, total_folders = self.last_result

        selected_filter = self.filter_segmented.get()
        search_text = self.search_var.get().strip().lower()

        filtered = folder_details

        if selected_filter != "Tümü":
            filtered = [item for item in filtered if item["category"] == selected_filter]

        if search_text:
            filtered = [
                item for item in filtered
                if search_text in item["folder"].lower()
            ]

        self.populate_tree(filtered)

        if selected_filter != "Tümü" or search_text:
            status_text = f"Filtre uygulandı"
            if search_text:
                status_text += f" | Arama: {self.search_var.get().strip()}"
            self.set_status(status_text, "Filtre")

    def on_search_change(self, event=None):
        self.apply_filters()

    def start_progress(self, indeterminate=True):
        self.progress.grid()
        self.progress.set(0)

        if indeterminate:
            self.progress.start()
        else:
            self.progress.stop()

    def stop_progress(self):
        self.progress.stop()
        self.progress.set(1)
        self.root.after(180, self.progress.grid_remove)

    def set_progress_value(self, value):
        self.progress.grid()
        self.progress.stop()
        self.progress.set(value)

    def clear_search(self):
        self.search_var.set("")
        self.search_entry.focus_set()
        self.apply_filters()

    def sort_treeview(self, column):
        if not self.last_result:
            return

        if self.sort_column == column:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = column
            self.sort_reverse = False

        self.apply_filters()

    def get_sorted_filtered_data(self, data):
        if not self.sort_column:
            return sorted(data, key=lambda x: x["count"], reverse=True)

        if self.sort_column == "folder":
            return sorted(
                data,
                key=lambda x: x["folder"].lower(),
                reverse=self.sort_reverse
            )

        if self.sort_column == "count":
            return sorted(
                data,
                key=lambda x: x["count"],
                reverse=self.sort_reverse
            )

        if self.sort_column == "category":
            order_map = {
                "0-5": 0,
                "6-25": 1,
                "26-50": 2,
                "50+": 3
            }
            return sorted(
                data,
                key=lambda x: order_map.get(x["category"], 999),
                reverse=self.sort_reverse
            )

        return data

    def update_tree_headings(self):
        folder_text = "Klasör"
        count_text = "Görüntü Sayısı"
        category_text = "Kategori"

        if self.sort_column == "folder":
            folder_text += " ↓" if self.sort_reverse else " ↑"
        elif self.sort_column == "count":
            count_text += " ↓" if self.sort_reverse else " ↑"
        elif self.sort_column == "category":
            category_text += " ↓" if self.sort_reverse else " ↑"

        self.tree.heading("folder", text=folder_text, command=lambda: self.sort_treeview("folder"))
        self.tree.heading("count", text=count_text, command=lambda: self.sort_treeview("count"))
        self.tree.heading("category", text=category_text, command=lambda: self.sort_treeview("category"))

    def update_tree_cursor(self, event):
        region = self.tree.identify_region(event.x, event.y)

        if region == "heading":
            self.tree.configure(cursor="hand2")
        else:
            self.tree.configure(cursor="")

    def on_analysis_mode_change(self, value):
        if value == "Standart":
            self.analysis_mode_info.configure(text="Sadece seçilen klasörün içindeki klasörleri sayar.")
        elif value == "Recursive":
            self.analysis_mode_info.configure(text="Seçilen klasörün alt klasörleri dahil tüm içeriğini sayar.")